#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
个人项目管理Excel模板生成器
根据设计文档创建包含12个工作表的完整Excel模板
"""

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule
from openpyxl.chart import PieChart, BarChart, LineChart, Reference
from datetime import datetime, timedelta
import os

class ExcelTemplateGenerator:
    def __init__(self):
        self.wb = Workbook()
        # 删除默认工作表
        self.wb.remove(self.wb.active)
        
        # 定义样式
        self.header_font = Font(name='微软雅黑', size=12, bold=True, color='FFFFFF')
        self.normal_font = Font(name='微软雅黑', size=10)
        self.header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        self.center_alignment = Alignment(horizontal='center', vertical='center')
        
        # 工作表配置
        self.worksheets_config = [
            ('📋需求收集', self.create_requirements_sheet),
            ('⭐需求优先级', self.create_priority_sheet),
            ('📊需求状态', self.create_status_sheet),
            ('📝任务分解', self.create_task_breakdown_sheet),
            ('⏱️工时估算', self.create_time_estimation_sheet),
            ('📈任务进度', self.create_task_progress_sheet),
            ('📅日程安排', self.create_schedule_sheet),
            ('🎯里程碑', self.create_milestone_sheet),
            ('⚠️风险评估', self.create_risk_assessment_sheet),
            ('📊统计报告', self.create_statistics_sheet),
            ('📚数据字典', self.create_data_dictionary_sheet),
            ('📖使用说明', self.create_instructions_sheet)
        ]
    
    def create_workbook(self):
        """创建完整的工作簿"""
        print("开始创建Excel项目管理模板...")
        
        for sheet_name, create_func in self.worksheets_config:
            print(f"创建工作表: {sheet_name}")
            ws = self.wb.create_sheet(title=sheet_name)
            create_func(ws)
        
        print("模板创建完成!")
        return self.wb
    
    def apply_header_style(self, ws, row, start_col, end_col):
        """应用表头样式"""
        for col in range(start_col, end_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.border
            cell.alignment = self.center_alignment
    
    def apply_data_style(self, ws, start_row, end_row, start_col, end_col):
        """应用数据区域样式"""
        for row in range(start_row, end_row + 1):
            for col in range(start_col, end_col + 1):
                cell = ws.cell(row=row, column=col)
                cell.font = self.normal_font
                cell.border = self.border
                cell.alignment = Alignment(horizontal='left', vertical='center')
    
    def create_requirements_sheet(self, ws):
        """创建需求收集表"""
        headers = [
            'A1:需求ID', 'B1:需求名称', 'C1:需求描述', 'D1:需求来源',
            'E1:提出人', 'F1:提出时间', 'G1:期望完成时间', 'H1:需求类型',
            'I1:业务价值', 'J1:技术复杂度', 'K1:备注'
        ]
        
        # 设置表头
        for i, header in enumerate(headers, 1):
            ws.cell(row=1, column=i, value=header.split(':')[1])
        
        self.apply_header_style(ws, 1, 1, 11)
        
        # 设置列宽
        column_widths = [12, 20, 30, 15, 12, 15, 15, 12, 12, 12, 20]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width
        
        # 添加数据验证
        # 需求来源下拉列表
        dv_source = DataValidation(type="list", formula1='"用户反馈,产品规划,技术优化,Bug修复,其他"')
        ws.add_data_validation(dv_source)
        dv_source.add(f'D2:D1000')
        
        # 需求类型下拉列表
        dv_type = DataValidation(type="list", formula1='"功能需求,性能需求,安全需求,易用性需求,兼容性需求"')
        ws.add_data_validation(dv_type)
        dv_type.add(f'H2:H1000')
        
        # 业务价值下拉列表
        dv_value = DataValidation(type="list", formula1='"高,中,低"')
        ws.add_data_validation(dv_value)
        dv_value.add(f'I2:I1000')
        
        # 技术复杂度下拉列表
        dv_complexity = DataValidation(type="list", formula1='"高,中,低"')
        ws.add_data_validation(dv_complexity)
        dv_complexity.add(f'J2:J1000')
        
        # 添加示例数据
        sample_data = [
            ['REQ-001', '用户登录功能', '实现用户账号登录和身份验证', '产品规划', '产品经理', '2024-01-15', '2024-02-15', '功能需求', '高', '中', '核心功能'],
            ['REQ-002', '数据导出功能', '支持Excel和CSV格式数据导出', '用户反馈', '用户A', '2024-01-16', '2024-02-20', '功能需求', '中', '低', ''],
            ['REQ-003', '性能优化', '提升系统响应速度', '技术优化', '技术负责人', '2024-01-17', '2024-03-01', '性能需求', '高', '高', '技术债务']
        ]
        
        for i, row_data in enumerate(sample_data, 2):
            for j, value in enumerate(row_data, 1):
                ws.cell(row=i, column=j, value=value)
        
        self.apply_data_style(ws, 2, 100, 1, 11)
        
        # 添加需求ID自动生成公式（从第5行开始）
        for row in range(5, 21):
            ws.cell(row=row, column=1, value=f'=IF(B{row}<>"","REQ-"&TEXT(ROW()-1,"000"),"")')
    
    def create_priority_sheet(self, ws):
        """创建需求优先级表"""
        headers = [
            'A1:需求ID', 'B1:需求名称', 'C1:业务价值(1-5)', 'D1:紧急程度(1-5)',
            'E1:实现难度(1-5)', 'F1:资源可用性(1-5)', 'G1:风险程度(1-5)',
            'H1:优先级得分', 'I1:最终优先级', 'J1:评估人', 'K1:评估时间'
        ]
        
        for i, header in enumerate(headers, 1):
            ws.cell(row=1, column=i, value=header.split(':')[1])
        
        self.apply_header_style(ws, 1, 1, 11)
        
        # 设置列宽
        column_widths = [12, 20, 15, 15, 15, 15, 15, 15, 15, 12, 15]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width
        
        # 添加数据验证（1-5分制）
        for col in ['C', 'D', 'E', 'F', 'G']:
            dv = DataValidation(type="list", formula1='"1,2,3,4,5"')
            ws.add_data_validation(dv)
            dv.add(f'{col}2:{col}1000')
        
        # 最终优先级下拉列表
        dv_priority = DataValidation(type="list", formula1='"P0-紧急,P1-高,P2-中,P3-低"')
        ws.add_data_validation(dv_priority)
        dv_priority.add(f'I2:I1000')
        
        # 添加示例数据和公式
        sample_data = [
            ['REQ-001', '用户登录功能', 5, 4, 3, 4, 2],
            ['REQ-002', '数据导出功能', 3, 3, 2, 4, 1],
            ['REQ-003', '性能优化', 4, 3, 4, 3, 3]
        ]
        
        for i, row_data in enumerate(sample_data, 2):
            for j, value in enumerate(row_data, 1):
                ws.cell(row=i, column=j, value=value)
            
            # 添加优先级得分计算公式
            ws.cell(row=i, column=8, value=f'=(C{i}*0.3+D{i}*0.25+E{i}*(-0.2)+F{i}*0.15+G{i}*(-0.1))')
            
            # 添加最终优先级判断公式
            ws.cell(row=i, column=9, value=f'=IF(H{i}>=4,"P0-紧急",IF(H{i}>=3,"P1-高",IF(H{i}>=2,"P2-中","P3-低")))')
            
            # 添加评估时间
            ws.cell(row=i, column=11, value=datetime.now().strftime('%Y-%m-%d'))
        
        self.apply_data_style(ws, 2, 100, 1, 11)
    
    def create_status_sheet(self, ws):
        """创建需求状态表"""
        headers = [
            'A1:需求ID', 'B1:需求名称', 'C1:当前状态', 'D1:进度(%)',
            'E1:负责人', 'F1:开始时间', 'G1:预计完成时间', 'H1:实际完成时间',
            'I1:状态更新时间', 'J1:遇到问题', 'K1:备注'
        ]
        
        for i, header in enumerate(headers, 1):
            ws.cell(row=1, column=i, value=header.split(':')[1])
        
        self.apply_header_style(ws, 1, 1, 11)
        
        # 设置列宽
        column_widths = [12, 20, 15, 10, 12, 15, 15, 15, 15, 20, 20]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width
        
        # 添加状态下拉列表
        dv_status = DataValidation(type="list", formula1='"待开始,进行中,已完成,已取消,阻塞,延期"')
        ws.add_data_validation(dv_status)
        dv_status.add(f'C2:C1000')
        
        # 添加示例数据
        sample_data = [
            ['REQ-001', '用户登录功能', '进行中', 60, '开发者A', '2024-01-20', '2024-02-15', '', '2024-01-25', '', '按计划进行'],
            ['REQ-002', '数据导出功能', '待开始', 0, '开发者B', '', '2024-02-20', '', '2024-01-16', '', '等待需求确认'],
            ['REQ-003', '性能优化', '已完成', 100, '开发者C', '2024-01-10', '2024-02-01', '2024-01-30', '2024-01-30', '', '提前完成']
        ]
        
        for i, row_data in enumerate(sample_data, 2):
            for j, value in enumerate(row_data, 1):
                ws.cell(row=i, column=j, value=value)
        
        self.apply_data_style(ws, 2, 100, 1, 11)
        
        # 添加条件格式
        # 进度条格式
        progress_rule = ColorScaleRule(
            start_type='num', start_value=0, start_color='FF6B6B',
            mid_type='num', mid_value=50, mid_color='FFE66D',
            end_type='num', end_value=100, end_color='4ECDC4'
        )
        ws.conditional_formatting.add('D2:D1000', progress_rule)
        
        # 状态颜色格式
        status_rules = [
            ('已完成', '4ECDC4'),
            ('进行中', '3498DB'),
            ('待开始', 'BDC3C7'),
            ('已取消', 'E74C3C'),
            ('阻塞', 'E74C3C'),
            ('延期', 'F39C12')
        ]
        
        for status, color in status_rules:
            rule = CellIsRule(operator='equal', formula=[f'"{status}"'], 
                            fill=PatternFill(start_color=color, end_color=color, fill_type='solid'))
            ws.conditional_formatting.add('C2:C1000', rule)
    
    def create_task_breakdown_sheet(self, ws):
        """创建任务分解表"""
        headers = [
            'A1:任务ID', 'B1:任务名称', 'C1:关联需求ID', 'D1:父任务ID',
            'E1:任务层级', 'F1:任务类型', 'G1:任务描述', 'H1:负责人',
            'I1:优先级', 'J1:依赖任务', 'K1:备注'
        ]
        
        for i, header in enumerate(headers, 1):
            ws.cell(row=1, column=i, value=header.split(':')[1])
        
        self.apply_header_style(ws, 1, 1, 11)
        
        # 设置列宽
        column_widths = [12, 25, 12, 12, 10, 12, 30, 12, 10, 15, 20]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width
        
        # 添加数据验证
        dv_type = DataValidation(type="list", formula1='"开发,测试,设计,文档,部署,其他"')
        ws.add_data_validation(dv_type)
        dv_type.add(f'F2:F1000')
        
        dv_priority = DataValidation(type="list", formula1='"高,中,低"')
        ws.add_data_validation(dv_priority)
        dv_priority.add(f'I2:I1000')
        
        # 添加示例数据
        sample_data = [
            ['TASK-001', '用户登录界面设计', 'REQ-001', '', 1, '设计', '设计登录页面UI', '设计师A', '高', '', ''],
            ['TASK-002', '登录API开发', 'REQ-001', '', 1, '开发', '实现用户认证API', '开发者A', '高', 'TASK-001', ''],
            ['TASK-003', '登录功能测试', 'REQ-001', '', 1, '测试', '测试登录功能', '测试员A', '中', 'TASK-002', ''],
            ['TASK-004', '数据导出界面', 'REQ-002', '', 1, '设计', '设计导出功能界面', '设计师A', '中', '', ''],
            ['TASK-005', '导出功能开发', 'REQ-002', '', 1, '开发', '实现数据导出功能', '开发者B', '中', 'TASK-004', '']
        ]
        
        for i, row_data in enumerate(sample_data, 2):
            for j, value in enumerate(row_data, 1):
                ws.cell(row=i, column=j, value=value)
        
        self.apply_data_style(ws, 2, 100, 1, 11)
        
        # 添加任务ID自动生成公式
        for row in range(7, 21):
            ws.cell(row=row, column=1, value=f'=IF(B{row}<>"","TASK-"&TEXT(ROW()-1,"000"),"")')
    
    def create_time_estimation_sheet(self, ws):
        """创建工时估算表"""
        headers = [
            'A1:任务ID', 'B1:任务名称', 'C1:乐观估算(小时)', 'D1:悲观估算(小时)',
            'E1:最可能估算(小时)', 'F1:期望工时', 'G1:复杂度系数', 'H1:最终估算',
            'I1:估算人', 'J1:估算时间', 'K1:备注'
        ]
        
        for i, header in enumerate(headers, 1):
            ws.cell(row=1, column=i, value=header.split(':')[1])
        
        self.apply_header_style(ws, 1, 1, 11)
        
        # 设置列宽
        column_widths = [12, 25, 15, 15, 15, 12, 12, 12, 12, 15, 20]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width
        
        # 添加复杂度系数下拉列表
        dv_complexity = DataValidation(type="list", formula1='"0.8,1.0,1.2,1.5,2.0"')
        ws.add_data_validation(dv_complexity)
        dv_complexity.add(f'G2:G1000')
        
        # 添加示例数据
        sample_data = [
            ['TASK-001', '用户登录界面设计', 4, 8, 6, '', 1.0, '', '估算员A', '2024-01-15'],
            ['TASK-002', '登录API开发', 8, 16, 12, '', 1.2, '', '估算员B', '2024-01-15'],
            ['TASK-003', '登录功能测试', 3, 6, 4, '', 1.0, '', '估算员A', '2024-01-15'],
            ['TASK-004', '数据导出界面', 3, 6, 4, '', 1.0, '', '估算员A', '2024-01-16'],
            ['TASK-005', '导出功能开发', 6, 12, 8, '', 1.2, '', '估算员B', '2024-01-16']
        ]
        
        for i, row_data in enumerate(sample_data, 2):
            for j, value in enumerate(row_data, 1):
                ws.cell(row=i, column=j, value=value)
            
            # 添加期望工时计算公式（三点估算法）
            ws.cell(row=i, column=6, value=f'=(C{i}+4*E{i}+D{i})/6')
            
            # 添加最终估算计算公式
            ws.cell(row=i, column=8, value=f'=F{i}*G{i}')
        
        self.apply_data_style(ws, 2, 100, 1, 11)
    
    def create_task_progress_sheet(self, ws):
        """创建任务进度表"""
        headers = [
            'A1:任务ID', 'B1:任务名称', 'C1:计划开始时间', 'D1:计划结束时间',
            'E1:实际开始时间', 'F1:实际结束时间', 'G1:进度(%)', 'H1:状态',
            'I1:预估工时', 'J1:实际工时', 'K1:工时偏差', 'L1:备注'
        ]
        
        for i, header in enumerate(headers, 1):
            ws.cell(row=1, column=i, value=header.split(':')[1])
        
        self.apply_header_style(ws, 1, 1, 12)
        
        # 设置列宽
        column_widths = [12, 25, 15, 15, 15, 15, 10, 12, 10, 10, 10, 20]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width
        
        # 添加状态下拉列表
        dv_status = DataValidation(type="list", formula1='"未开始,进行中,已完成,已取消,阻塞,延期"')
        ws.add_data_validation(dv_status)
        dv_status.add(f'H2:H1000')
        
        # 添加示例数据
        sample_data = [
            ['TASK-001', '用户登录界面设计', '2024-01-20', '2024-01-22', '2024-01-20', '2024-01-21', 100, '已完成', 6, 5, '', '提前完成'],
            ['TASK-002', '登录API开发', '2024-01-22', '2024-01-26', '2024-01-22', '', 75, '进行中', 14.4, 12, '', '按计划进行'],
            ['TASK-003', '登录功能测试', '2024-01-26', '2024-01-28', '', '', 0, '未开始', 4, 0, '', '等待开发完成'],
            ['TASK-004', '数据导出界面', '2024-01-25', '2024-01-27', '', '', 0, '未开始', 4, 0, '', ''],
            ['TASK-005', '导出功能开发', '2024-01-27', '2024-01-31', '', '', 0, '未开始', 9.6, 0, '', '']
        ]
        
        for i, row_data in enumerate(sample_data, 2):
            for j, value in enumerate(row_data, 1):
                ws.cell(row=i, column=j, value=value)
            
            # 添加工时偏差计算公式
            ws.cell(row=i, column=11, value=f'=IF(J{i}>0,(J{i}-I{i})/I{i},"")')
        
        self.apply_data_style(ws, 2, 100, 1, 12)
        
        # 添加条件格式
        # 进度条格式
        progress_rule = ColorScaleRule(
            start_type='num', start_value=0, start_color='FF6B6B',
            mid_type='num', mid_value=50, mid_color='FFE66D',
            end_type='num', end_value=100, end_color='4ECDC4'
        )
        ws.conditional_formatting.add('G2:G1000', progress_rule)
        
        # 工时偏差预警
        deviation_rule = CellIsRule(operator='greaterThan', formula=[0.2], 
                                  fill=PatternFill(start_color='F39C12', end_color='F39C12', fill_type='solid'))
        ws.conditional_formatting.add('K2:K1000', deviation_rule)
    
    def create_schedule_sheet(self, ws):
        """创建日程安排表"""
        headers = [
            'A1:日期', 'B1:时间段', 'C1:关联任务ID', 'D1:工作内容',
            'E1:计划工时', 'F1:实际工时', 'G1:完成情况', 'H1:工作地点',
            'I1:参与人员', 'J1:备注'
        ]
        
        for i, header in enumerate(headers, 1):
            ws.cell(row=1, column=i, value=header.split(':')[1])
        
        self.apply_header_style(ws, 1, 1, 10)
        
        # 设置列宽
        column_widths = [12, 15, 12, 30, 10, 10, 12, 12, 15, 20]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width
        
        # 添加时间段下拉列表
        dv_time = DataValidation(type="list", formula1='"上午,下午,晚上,全天"')
        ws.add_data_validation(dv_time)
        dv_time.add(f'B2:B1000')
        
        # 添加完成情况下拉列表
        dv_completion = DataValidation(type="list", formula1='"已完成,进行中,未开始,取消"')
        ws.add_data_validation(dv_completion)
        dv_completion.add(f'G2:G1000')
        
        # 添加工作地点下拉列表
        dv_location = DataValidation(type="list", formula1='"办公室,家里,客户现场,其他"')
        ws.add_data_validation(dv_location)
        dv_location.add(f'H2:H1000')
        
        # 添加示例数据
        today = datetime.now()
        sample_data = [
            [today.strftime('%Y-%m-%d'), '上午', 'TASK-001', '完成登录界面设计', 4, 3, '已完成', '办公室', '设计师A', ''],
            [(today + timedelta(days=1)).strftime('%Y-%m-%d'), '全天', 'TASK-002', '开发登录API', 8, 6, '进行中', '家里', '开发者A', ''],
            [(today + timedelta(days=2)).strftime('%Y-%m-%d'), '上午', 'TASK-003', '测试登录功能', 4, 0, '未开始', '办公室', '测试员A', ''],
            [(today + timedelta(days=3)).strftime('%Y-%m-%d'), '下午', 'TASK-004', '设计导出界面', 4, 0, '未开始', '办公室', '设计师A', ''],
            [(today + timedelta(days=4)).strftime('%Y-%m-%d'), '全天', 'TASK-005', '开发导出功能', 8, 0, '未开始', '家里', '开发者B', '']
        ]
        
        for i, row_data in enumerate(sample_data, 2):
            for j, value in enumerate(row_data, 1):
                ws.cell(row=i, column=j, value=value)
        
        self.apply_data_style(ws, 2, 100, 1, 10)
    
    def create_milestone_sheet(self, ws):
        """创建里程碑表"""
        headers = [
            'A1:里程碑ID', 'B1:里程碑名称', 'C1:描述', 'D1:计划完成时间',
            'E1:实际完成时间', 'F1:状态', 'G1:完成标准', 'H1:验收人',
            'I1:关联需求', 'J1:备注'
        ]
        
        for i, header in enumerate(headers, 1):
            ws.cell(row=1, column=i, value=header.split(':')[1])
        
        self.apply_header_style(ws, 1, 1, 10)
        
        # 设置列宽
        column_widths = [12, 20, 30, 15, 15, 12, 25, 12, 15, 20]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width
        
        # 添加状态下拉列表
        dv_status = DataValidation(type="list", formula1='"未开始,进行中,已完成,延期"')
        ws.add_data_validation(dv_status)
        dv_status.add(f'F2:F1000')
        
        # 添加示例数据
        sample_data = [
            ['MS-001', '用户认证模块完成', '完成用户登录、注册、权限验证功能', '2024-02-15', '2024-02-10', '已完成', '所有登录相关功能测试通过', '产品经理', 'REQ-001', '提前完成'],
            ['MS-002', '数据管理模块完成', '完成数据导入、导出、查询功能', '2024-03-01', '', '进行中', '数据操作功能测试通过', '产品经理', 'REQ-002', ''],
            ['MS-003', '系统性能优化完成', '完成系统性能优化和监控', '2024-03-15', '', '未开始', '性能指标达到预期', '技术负责人', 'REQ-003', ''],
            ['MS-004', 'MVP版本发布', '完成最小可行产品版本', '2024-04-01', '', '未开始', '核心功能可用', '项目经理', 'REQ-001,REQ-002', '']
        ]
        
        for i, row_data in enumerate(sample_data, 2):
            for j, value in enumerate(row_data, 1):
                ws.cell(row=i, column=j, value=value)
        
        self.apply_data_style(ws, 2, 100, 1, 10)
        
        # 添加里程碑ID自动生成公式
        for row in range(6, 21):
            ws.cell(row=row, column=1, value=f'=IF(B{row}<>"","MS-"&TEXT(ROW()-1,"000"),"")')
    
    def create_risk_assessment_sheet(self, ws):
        """创建风险评估表"""
        headers = [
            'A1:风险ID', 'B1:风险名称', 'C1:风险描述', 'D1:风险类型',
            'E1:发生概率', 'F1:影响程度', 'G1:风险值', 'H1:风险等级',
            'I1:预防措施', 'J1:应对措施', 'K1:负责人', 'L1:状态'
        ]
        
        for i, header in enumerate(headers, 1):
            ws.cell(row=1, column=i, value=header.split(':')[1])
        
        self.apply_header_style(ws, 1, 1, 12)
        
        # 设置列宽
        column_widths = [12, 20, 25, 12, 10, 10, 10, 12, 25, 25, 12, 10]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width
        
        # 添加数据验证
        dv_type = DataValidation(type="list", formula1='"技术风险,进度风险,资源风险,质量风险,外部风险"')
        ws.add_data_validation(dv_type)
        dv_type.add(f'D2:D1000')
        
        dv_probability = DataValidation(type="list", formula1='"1,2,3,4,5"')
        ws.add_data_validation(dv_probability)
        dv_probability.add(f'E2:E1000')
        
        dv_impact = DataValidation(type="list", formula1='"1,2,3,4,5"')
        ws.add_data_validation(dv_impact)
        dv_impact.add(f'F2:F1000')
        
        dv_status = DataValidation(type="list", formula1='"识别,监控,发生,已解决"')
        ws.add_data_validation(dv_status)
        dv_status.add(f'L2:L1000')
        
        # 添加示例数据
        sample_data = [
            ['RISK-001', '技术难度超预期', 'API开发复杂度可能超出预期', '技术风险', 3, 4, '', '', '提前技术调研', '寻求技术支持', '技术负责人', '监控'],
            ['RISK-002', '第三方服务不稳定', '依赖的第三方服务可能不稳定', '外部风险', 2, 3, '', '', '准备备用方案', '切换备用服务', '架构师', '监控'],
            ['RISK-003', '人员离职', '关键开发人员可能离职', '资源风险', 2, 5, '', '', '知识文档化', '人员备份培训', 'HR', '监控'],
            ['RISK-004', '需求变更频繁', '客户需求可能频繁变更', '进度风险', 4, 3, '', '', '需求冻结机制', '变更控制流程', '产品经理', '监控']
        ]
        
        for i, row_data in enumerate(sample_data, 2):
            for j, value in enumerate(row_data, 1):
                ws.cell(row=i, column=j, value=value)
            
            # 添加风险值计算公式
            ws.cell(row=i, column=7, value=f'=E{i}*F{i}')
            
            # 添加风险等级判断公式
            ws.cell(row=i, column=8, value=f'=IF(G{i}>=15,"高风险",IF(G{i}>=9,"中风险","低风险"))')
        
        self.apply_data_style(ws, 2, 100, 1, 12)
        
        # 添加风险等级条件格式
        high_risk_rule = CellIsRule(operator='equal', formula=['"高风险"'], 
                                  fill=PatternFill(start_color='E74C3C', end_color='E74C3C', fill_type='solid'))
        ws.conditional_formatting.add('H2:H1000', high_risk_rule)
        
        medium_risk_rule = CellIsRule(operator='equal', formula=['"中风险"'], 
                                    fill=PatternFill(start_color='F39C12', end_color='F39C12', fill_type='solid'))
        ws.conditional_formatting.add('H2:H1000', medium_risk_rule)
        
        low_risk_rule = CellIsRule(operator='equal', formula=['"低风险"'], 
                                 fill=PatternFill(start_color='27AE60', end_color='27AE60', fill_type='solid'))
        ws.conditional_formatting.add('H2:H1000', low_risk_rule)
    
    def create_statistics_sheet(self, ws):
        """创建统计报告表"""
        # 项目基本信息
        ws.cell(row=1, column=1, value='项目统计报告')
        ws.cell(row=1, column=1).font = Font(name='微软雅黑', size=16, bold=True)
        ws.merge_cells('A1:F1')
        
        # 统计指标
        stats_data = [
            ['统计项目', '数值', '单位', '说明'],
            ['需求总数', '=COUNTA(需求收集.B:B)-1', '个', '已收集的需求总数'],
            ['已完成需求', '=COUNTIF(需求状态.C:C,"已完成")', '个', '已完成的需求数量'],
            ['需求完成率', '=IF(B3>0,B4/B3,0)', '%', '需求完成百分比'],
            ['任务总数', '=COUNTA(任务分解.B:B)-1', '个', '已分解的任务总数'],
            ['已完成任务', '=COUNTIF(任务进度.H:H,"已完成")', '个', '已完成的任务数量'],
            ['任务完成率', '=IF(B6>0,B7/B6,0)', '%', '任务完成百分比'],
            ['总预估工时', '=SUM(工时估算.H:H)', '小时', '所有任务预估工时总和'],
            ['总实际工时', '=SUM(任务进度.J:J)', '小时', '所有任务实际工时总和'],
            ['工时偏差率', '=IF(B9>0,(B10-B9)/B9,0)', '%', '实际工时与预估工时的偏差'],
            ['高风险数量', '=COUNTIF(风险评估.H:H,"高风险")', '个', '高风险项目数量'],
            ['里程碑总数', '=COUNTA(里程碑.B:B)-1', '个', '设定的里程碑总数'],
            ['已完成里程碑', '=COUNTIF(里程碑.F:F,"已完成")', '个', '已完成的里程碑数量']
        ]
        
        for i, row_data in enumerate(stats_data, 3):
            for j, value in enumerate(row_data, 1):
                ws.cell(row=i+2, column=j, value=value)
        
        # 应用样式
        self.apply_header_style(ws, 5, 1, 4)
        self.apply_data_style(ws, 6, 20, 1, 4)
        
        # 设置列宽
        column_widths = [15, 12, 8, 25]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width
        
        # 添加图表区域说明
        ws.cell(row=22, column=1, value='图表区域')
        ws.cell(row=22, column=1).font = Font(name='微软雅黑', size=14, bold=True)
        
        # 创建需求状态分布饼图
        pie_chart = PieChart()
        pie_chart.title = "需求状态分布"
        pie_chart.height = 10
        pie_chart.width = 15
        
        # 图表数据（这里使用示例数据，实际应该引用其他工作表）
        chart_data = Reference(ws, min_col=1, min_row=24, max_row=27, max_col=2)
        pie_chart.add_data(chart_data, titles_from_data=True)
        
        # 添加图表数据
        chart_sample_data = [
            ['状态', '数量'],
            ['已完成', 1],
            ['进行中', 1], 
            ['待开始', 1]
        ]
        
        for i, row_data in enumerate(chart_sample_data, 24):
            for j, value in enumerate(row_data, 1):
                ws.cell(row=i, column=j, value=value)
        
        ws.add_chart(pie_chart, "D24")
    
    def create_data_dictionary_sheet(self, ws):
        """创建数据字典表"""
        ws.cell(row=1, column=1, value='数据字典')
        ws.cell(row=1, column=1).font = Font(name='微软雅黑', size=16, bold=True)
        ws.merge_cells('A1:D1')
        
        # 字典数据
        dictionary_data = [
            ['类别', '选项值', '显示名称', '说明'],
            ['需求来源', 'user_feedback', '用户反馈', '来自用户的反馈和建议'],
            ['需求来源', 'product_planning', '产品规划', '产品规划阶段提出的需求'],
            ['需求来源', 'tech_optimization', '技术优化', '技术团队提出的优化需求'],
            ['需求来源', 'bug_fix', 'Bug修复', '修复系统缺陷的需求'],
            ['需求来源', 'other', '其他', '其他来源的需求'],
            ['需求类型', 'functional', '功能需求', '系统功能相关需求'],
            ['需求类型', 'performance', '性能需求', '系统性能相关需求'],
            ['需求类型', 'security', '安全需求', '系统安全相关需求'],
            ['需求类型', 'usability', '易用性需求', '用户体验相关需求'],
            ['需求类型', 'compatibility', '兼容性需求', '系统兼容性相关需求'],
            ['优先级', 'P0', 'P0-紧急', '最高优先级，需要立即处理'],
            ['优先级', 'P1', 'P1-高', '高优先级，需要优先处理'],
            ['优先级', 'P2', 'P2-中', '中等优先级，按计划处理'],
            ['优先级', 'P3', 'P3-低', '低优先级，有时间再处理'],
            ['状态', 'pending', '待开始', '任务尚未开始'],
            ['状态', 'in_progress', '进行中', '任务正在进行'],
            ['状态', 'completed', '已完成', '任务已经完成'],
            ['状态', 'cancelled', '已取消', '任务被取消'],
            ['状态', 'blocked', '阻塞', '任务被阻塞无法进行'],
            ['状态', 'delayed', '延期', '任务延期处理'],
            ['任务类型', 'development', '开发', '代码开发任务'],
            ['任务类型', 'testing', '测试', '测试相关任务'],
            ['任务类型', 'design', '设计', '设计相关任务'],
            ['任务类型', 'documentation', '文档', '文档编写任务'],
            ['任务类型', 'deployment', '部署', '部署相关任务'],
            ['任务类型', 'other', '其他', '其他类型任务'],
            ['风险类型', 'technical', '技术风险', '技术实现相关风险'],
            ['风险类型', 'schedule', '进度风险', '项目进度相关风险'],
            ['风险类型', 'resource', '资源风险', '人力资源相关风险'],
            ['风险类型', 'quality', '质量风险', '产品质量相关风险'],
            ['风险类型', 'external', '外部风险', '外部环境相关风险']
        ]
        
        for i, row_data in enumerate(dictionary_data, 3):
            for j, value in enumerate(row_data, 1):
                ws.cell(row=i, column=j, value=value)
        
        # 应用样式
        self.apply_header_style(ws, 3, 1, 4)
        self.apply_data_style(ws, 4, 35, 1, 4)
        
        # 设置列宽
        column_widths = [15, 20, 15, 30]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width
    
    def create_instructions_sheet(self, ws):
        """创建使用说明表"""
        ws.cell(row=1, column=1, value='Excel项目管理模板使用说明')
        ws.cell(row=1, column=1).font = Font(name='微软雅黑', size=16, bold=True)
        ws.merge_cells('A1:F1')
        
        instructions = [
            ['', ''],
            ['一、模板概述', ''],
            ['本模板为个人项目管理提供完整解决方案，包含需求管理、任务分解、进度跟踪、风险评估等功能。', ''],
            ['', ''],
            ['二、工作表说明', ''],
            ['📋需求收集', '记录和管理项目需求'],
            ['⭐需求优先级', '评估需求优先级'],
            ['📊需求状态', '跟踪需求状态'],
            ['📝任务分解', '将需求分解为具体任务'],
            ['⏱️工时估算', '估算任务工时'],
            ['📈任务进度', '跟踪任务执行进度'],
            ['📅日程安排', '安排日常工作计划'],
            ['🎯里程碑', '设定和跟踪项目里程碑'],
            ['⚠️风险评估', '识别和管理项目风险'],
            ['📊统计报告', '查看项目统计数据'],
            ['📚数据字典', '查看下拉选项说明'],
            ['📖使用说明', '本使用说明'],
            ['', ''],
            ['三、使用流程', ''],
            ['1. 需求管理', '在需求收集表中添加需求 → 在需求优先级表中评估 → 在需求状态表中跟踪'],
            ['2. 任务管理', '在任务分解表中分解需求 → 在工时估算表中估算 → 在任务进度表中跟踪'],
            ['3. 计划管理', '在日程安排表中安排工作 → 在里程碑表中设定目标 → 在风险评估表中管理风险'],
            ['4. 数据分析', '在统计报告表中查看项目整体情况'],
            ['', ''],
            ['四、注意事项', ''],
            ['1. 请勿删除表头行，避免公式错误', ''],
            ['2. 使用下拉列表选择选项，保持数据一致性', ''],
            ['3. 定期更新进度和状态，确保数据准确性', ''],
            ['4. 备份文件，避免数据丢失', ''],
            ['5. 可根据项目特点调整字段和选项', ''],
            ['', ''],
            ['五、公式说明', ''],
            ['需求ID自动生成', '=IF(B2<>"","REQ-"&TEXT(ROW()-1,"000"),"")'],
            ['优先级得分计算', '=(C2*0.3+D2*0.25+E2*(-0.2)+F2*0.15+G2*(-0.1))'],
            ['期望工时计算', '=(C2+4*E2+D2)/6'],
            ['工时偏差计算', '=(实际工时-预估工时)/预估工时'],
            ['风险值计算', '=发生概率*影响程度'],
            ['', ''],
            ['六、技术支持', ''],
            ['如有问题，请参考数据字典或联系技术支持。', ''],
            ['版本：v1.0', ''],
            ['更新日期：2024年1月', '']
        ]
        
        for i, (col1, col2) in enumerate(instructions, 3):
            ws.cell(row=i, column=1, value=col1)
            ws.cell(row=i, column=2, value=col2)
            
            # 设置标题样式
            if col1.startswith(('一、', '二、', '三、', '四、', '五、', '六、')):
                ws.cell(row=i, column=1).font = Font(name='微软雅黑', size=12, bold=True)
            elif col1 and not col1.startswith(' '):
                ws.cell(row=i, column=1).font = Font(name='微软雅黑', size=10, bold=True)
        
        # 设置列宽
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 40
    
    def save_template(self, filename='个人项目管理模板.xlsx'):
        """保存模板文件"""
        filepath = os.path.join('/Users/zhuwencan/work/phoenixcoder', filename)
        self.wb.save(filepath)
        print(f"模板已保存到: {filepath}")
        return filepath

def main():
    """主函数"""
    print("开始生成Excel项目管理模板...")
    
    generator = ExcelTemplateGenerator()
    generator.create_workbook()
    filepath = generator.save_template()
    
    print("\n模板生成完成！")
    print(f"文件位置: {filepath}")
    print("\n模板包含以下功能:")
    print("✅ 12个专业工作表")
    print("✅ 完整的数据验证规则")
    print("✅ 自动计算公式")
    print("✅ 条件格式和颜色标识")
    print("✅ 示例数据和使用说明")
    print("\n请打开Excel文件开始使用！")

if __name__ == '__main__':
    main()